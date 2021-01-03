from random import choice,randrange
from string import digits,ascii_letters
from os import listdir,mkdir
from os.path import isdir
import sqlite3
from time import time
from openpyxl import Workbook,load_workbook
def generateRandomData():
    if not isdir('xlsxs'):
        mkdir('xlsxs')
    global total
    characters=digits+ascii_letters
    for i in range(3):
        xlsName='xlsxs\\'+str(i)+'.xlsx'
        totalLines=randrange(10)
        wb=Workbook()
        ws=wb.worksheets[0]
        ws.append(['a','b','c'])
        for j in range(totalLines):
            line=[''.join((choice(characters)for ii in range(30)))
                  for jj in range(5)]
            ws.append(line)
            total+=1
        wb.save(xlsName)
def eachXlsx(xlsxFn):
    wb=load_workbook(xlsxFn)
    ws=wb.worksheets[0]
    for index,row in enumerate(ws.rows):
        if index==0:
            continue
        yield tuple(map(lambda x:x.value,row))
def xlsx2sqlite():
    xlsxs=('xlsxs\\'+fn for fn in listdir('xlsxs'))
    with sqlite3.connect('dataxlsx.db') as conn:
        #print('eee')
        cur=conn.cursor()
        for xlsx in xlsxs:
            sql='INSERT INTO fromxlsx VALUES(?,?,?,?,?)'
            cur.executemany(sql,eachXlsx(xlsx))
            conn.commit()
total=0
generateRandomData()
start=time()
xlsx2sqlite()
delta=time()-start
print('导入用时:',delta)
print('导入速度(条/秒):',total/delta)
